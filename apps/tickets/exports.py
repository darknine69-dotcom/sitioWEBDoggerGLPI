from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone

HEADER_FILL = PatternFill(start_color="1D3557", end_color="1D3557", fill_type="solid")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="D62B1F")

COLUMNS = [
    ("Código", "codigo", 12),
    ("Título", "titulo", 34),
    ("Solicitante", "solicitante_nombre", 22),
    ("Correo", "solicitante_email", 26),
    ("Punto / equipo", "solicitante_punto", 22),
    ("Categoría", "_categoria", 20),
    ("Prioridad", "_prioridad", 12),
    ("Estado", "_estado", 14),
    ("Técnico asignado", "_tecnico", 20),
    ("GLPI #", "glpi_id", 8),
    ("Fecha creación", "_fecha_creacion", 18),
    ("Fecha cierre", "_fecha_cierre", 18),
]


def _row_values(ticket):
    return [
        ticket.codigo,
        ticket.titulo,
        ticket.solicitante_nombre,
        ticket.solicitante_email or "",
        ticket.solicitante_punto or "",
        ticket.categoria.nombre if ticket.categoria else "Sin categoría",
        ticket.get_prioridad_display(),
        ticket.get_estado_display(),
        ticket.tecnico_asignado.nombre if ticket.tecnico_asignado else "",
        ticket.glpi_id or "",
        timezone.localtime(ticket.fecha_creacion).strftime("%d/%m/%Y %H:%M"),
        timezone.localtime(ticket.fecha_cierre).strftime("%d/%m/%Y %H:%M") if ticket.fecha_cierre else "",
    ]


def generar_excel_tickets(tickets, titulo="Tickets Dogger Helpdesk"):
    """
    Recibe un queryset/lista de Ticket (con select_related aplicado por el
    caller) y retorna un HttpResponse .xlsx listo para descargar.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"

    ws.merge_cells("A1:L1")
    ws["A1"] = titulo
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Generado el {timezone.localtime(timezone.now()).strftime('%d/%m/%Y %H:%M')}"
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")

    header_row = 4
    for col_idx, (label, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for r, ticket in enumerate(tickets, start=header_row + 1):
        values = _row_values(ticket)
        for c, value in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            cell.font = BODY_FONT

    ws.freeze_panes = f"A{header_row + 1}"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"dogger_tickets_{timezone.localtime(timezone.now()).strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
